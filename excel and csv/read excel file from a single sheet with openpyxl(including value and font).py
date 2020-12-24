import openpyxl

# Read excel file
input_dt = {}
input_fname = 'My Excel.xlsx'

wb_obj = openpyxl.load_workbook(input_fname)
sheet_names = wb_obj.sheetnames

for i, sheet_name in enumerate(sheet_names):
    sheet = wb_obj[sheet_name]

    if sheet_name not in input_dt:
        input_dt[sheet_name] = []

    for row_index in range(0, sheet.max_row):
        row = [(sheet.cell(row_index+1, col_index+1).value, sheet.cell(row_index+1, col_index+1).font) for col_index in range(sheet.max_column)]
        input_dt[sheet_name].append(row)