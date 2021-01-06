import openpyxl

wb = openpyxl.Workbook()
sheet_2021 = wb.create_sheet("2021")
sheet_2022 = wb.create_sheet("2022")
sheet_2023 = wb.create_sheet("2023")
sheet_2024 = wb.create_sheet("2024")
sheet_2025 = wb.create_sheet("2025")

wb.save("Temp.xlsx")


def insert_row(sheet, sheet_cur_idx, result_row):
    for i, elm in enumerate(result_row):
        sheet.cell(row=sheet_cur_idx, column=i + 1).value = elm